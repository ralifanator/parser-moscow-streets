import requests
import os
import csv
from fake_useragent import *
import openpyxl as pyxl
from tqdm import tqdm
import time
import random

from get_inf_about_street import *
from get_streets_links import *

file = 'streets_urls.txt'
file_excel = 'streets.xlsx'
wb = pyxl.Workbook()
sheet = wb.active

ua = UserAgent()
headers = {'User-Agent': ua.chrome}

if not os.path.isfile(file):
    write_streets_links(parse_streets(), 'streets_urls.txt')

streets_name = []
streets_links = []
result = {}
i = 1
with open(file, 'r') as f:
    f_reader = csv.reader(f, delimiter=';')
    sheet.cell(row=i, column=1).value = 'Название'
    sheet.cell(row=i, column=2).value = 'Район'
    sheet.cell(row=i, column=3).value = 'Дома'
    sheet.cell(row=i, column=4).value = 'Перекрестки'
    sheet.cell(row=i, column=5).value = 'URL'
    i += 1
    progress_bar = tqdm(f_reader)
    for row in progress_bar:
        url = 'https://ginfo.ru' + row[1]
        req = requests.get(url, headers=headers)
        full_inf = {}
        
        full_inf['name'] = row[0]
        full_inf['url'] = url

        symbols_remove = "'[]"    
        full_inf['district'] = get_district(req)
        sheet.cell(row=i, column=1).value = row[0]
        
        sheet.cell(row=i, column=2).value = clean_string(str(full_inf['district']))
        full_inf['houses'] = get_house_numbers(req)
        sheet.cell(row=i, column=3).value = clean_string(str(full_inf['houses']))
        full_inf['crossroads'] = get_crossroads(req)
        sheet.cell(row=i, column=4).value = clean_string(str(full_inf['crossroads']))
        sheet.cell(row=i, column=5).value = url

        i+=1
        
        time.sleep(random.randint(2, 10))
        tqdm.write(row[1])
    
        wb.save("moscow_streets.xlsx")
            
            
        
            


